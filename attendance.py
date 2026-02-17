import cv2
import openpyxl
import os
from datetime import datetime, timedelta
from tkinter import Tk, messagebox, Label, Button


class FaceRecognitionSystem:
    def __init__(self, model_path: str, excel_path: str):
        self.model_path = model_path
        self.excel_file = excel_path

        self.clf = cv2.face.LBPHFaceRecognizer_create()
        self.read_model()

        self.init_excel_file()
        self.student_data = self.load_student_data()

        self.face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
        )

        self.recognized_face_id = None
        self.recognized_face_name = None
        self.root = None

    # ---------------- MODEL ---------------- #

    def read_model(self):
        if not os.path.exists(self.model_path):
            raise FileNotFoundError(f"Model file not found: {self.model_path}")
        self.clf.read(self.model_path)

    # ---------------- EXCEL ---------------- #

    def init_excel_file(self):
        if not os.path.exists(self.excel_file):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Student Data"
            sheet.append([
                "ID", "Name", "Division", "Gender",
                "DOB", "Email", "Phone No",
                "Address", "Teacher", "PhotoSample"
            ])
            wb.save(self.excel_file)

    def load_student_data(self):
        wb = openpyxl.load_workbook(self.excel_file)
        sheet = wb.active
        data = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            try:
                student_id = int(row[0])
                name = str(row[1])
                data[student_id] = name
            except:
                continue

        return data

    # ---------------- ATTENDANCE FILE ---------------- #

    def create_attendance_sheet(self):
        file_name = "attendance.xlsx"

        if not os.path.exists(file_name):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "Attendance"
            sheet.append(["ID", "Name", "Date", "Period", "Time"])
            wb.save(file_name)

        return file_name

    def get_current_period(self, time_str):
        hour = int(time_str.split(":")[0])
        return f"{hour + 1}th Period"

    # ---------------- FACE RECOGNITION ---------------- #

    def recognize_faces(self):

        self.recognized_face_id = None
        self.recognized_face_name = None

        cap = cv2.VideoCapture(0)

        if not cap.isOpened():
            messagebox.showerror("Error", "Camera not accessible")
            return

        start_time = datetime.now()

        # 🔥 Detection time reduced to 8 seconds
        while (datetime.now() - start_time) < timedelta(seconds=8):

            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

            # ⚡ Faster detection parameters
            faces = self.face_cascade.detectMultiScale(
                gray,
                scaleFactor=1.2,
                minNeighbors=5,
                minSize=(80, 80)
            )

            for (x, y, w, h) in faces:
                face = gray[y:y + h, x:x + w]
                face = cv2.resize(face, (200, 200))

                try:
                    student_id, confidence = self.clf.predict(face)
                except cv2.error:
                    continue

                name = self.student_data.get(student_id, "Unknown")

                if confidence < 60:
                    self.recognized_face_id = student_id
                    self.recognized_face_name = name

                    cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                    cv2.putText(
                        frame,
                        f"{name} ({round(100 - confidence)}%)",
                        (x, y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.8,
                        (0, 255, 0),
                        2,
                    )

            cv2.imshow("Face Recognition", frame)

            # Press Enter to confirm early
            if cv2.waitKey(1) & 0xFF == 13:
                break

        cap.release()
        cv2.destroyAllWindows()

        if self.recognized_face_id:
            self.show_confirmation_dialog()
        else:
            messagebox.showinfo("Info", "Face not recognized. Try again.")

    # ---------------- CONFIRMATION ---------------- #

    def show_confirmation_dialog(self):
        result = messagebox.askyesno(
            "Confirm Attendance",
            f"Is this {self.recognized_face_name}?"
        )

        if result:
            self.log_attendance()

    # ---------------- LOG ATTENDANCE ---------------- #

    def log_attendance(self):
        file_name = self.create_attendance_sheet()
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active

        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")
        period = self.get_current_period(now.strftime("%H:%M"))

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == self.recognized_face_id and row[2] == date:
                messagebox.showinfo("Info", "Attendance already recorded today.")
                return

        sheet.append([
            self.recognized_face_id,
            self.recognized_face_name,
            date,
            period,
            time
        ])

        wb.save(file_name)
        messagebox.showinfo("Success", "Attendance Logged Successfully")
        self.quit_gui()

    # ---------------- GUI ---------------- #

    def run_gui(self):
        self.root = Tk()
        self.root.title("Face Recognition Attendance System")
        self.root.geometry("400x250")

        Label(
            self.root,
            text="Face Recognition Attendance System",
            font=("Arial", 14)
        ).pack(pady=30)

        Button(
            self.root,
            text="Take Attendance",
            command=self.recognize_faces,
            font=("Arial", 12),
            width=20
        ).pack(pady=20)

        self.root.mainloop()

    def quit_gui(self):
        if self.root:
            self.root.quit()
            self.root.destroy()


# ---------------- MAIN ---------------- #

if __name__ == "__main__":
    model_path = "trainer.yml"
    excel_path = "student_data.xlsx"

    try:
        app = FaceRecognitionSystem(model_path, excel_path)
        app.run_gui()
    except Exception as e:
        print(f"Error: {e}")
