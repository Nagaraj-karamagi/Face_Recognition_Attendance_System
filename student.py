import tkinter as tk
from tkinter import ttk, Label, LabelFrame, Frame, Button, StringVar, messagebox
from PIL import Image, ImageTk
from tkinter.constants import RIDGE, W
from openpyxl import Workbook, load_workbook
import os
import cv2


class FaceRecognitionSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Face Recognition System")
        self.root.geometry('1261x843+0+0')

        # Variables
        self.name = StringVar()
        self.division = StringVar()
        self.id = StringVar()
        self.gender = StringVar()
        self.dob = StringVar()
        self.email = StringVar()
        self.phone_no = StringVar()
        self.address = StringVar()
        self.teacher = StringVar()

        self.excel_file = "student_data.xlsx"
        self.init_excel_file()

        self.setup_ui()
        self.load_data()

    # ---------------- UI ---------------- #

    def setup_ui(self):

        left_frame = LabelFrame(self.root, bd=2, bg='white',
                                relief=RIDGE, text="Student Details")
        left_frame.place(x=5, y=5, width=620, height=400)

        labels = [
            ("Student ID", self.id),
            ("Student Name", self.name),
            ("Division", self.division),
            ("Gender", self.gender),
            ("DOB", self.dob),
            ("Email", self.email),
            ("Phone No", self.phone_no),
            ("Address", self.address),
            ("Teacher", self.teacher)
        ]

        for i, (text, var) in enumerate(labels):
            Label(left_frame, text=text).grid(row=i, column=0, padx=10, pady=5, sticky=W)

            if text == "Gender":
                ttk.Combobox(left_frame, textvariable=var,
                             values=["Male", "Female", "Other"],
                             state="readonly").grid(row=i, column=1)
            else:
                ttk.Entry(left_frame, textvariable=var).grid(row=i, column=1)

        btn_frame = Frame(left_frame)
        btn_frame.place(x=0, y=280, width=600)

        Button(btn_frame, text="Save", command=self.add_data).grid(row=0, column=0, padx=5)
        Button(btn_frame, text="Update", command=self.update_data).grid(row=0, column=1, padx=5)
        Button(btn_frame, text="Delete", command=self.delete_data).grid(row=0, column=2, padx=5)
        Button(btn_frame, text="Reset", command=self.reset_data).grid(row=0, column=3, padx=5)
        Button(btn_frame, text="Capture Face", command=self.take_photo_sample).grid(row=0, column=4, padx=5)

        # Table
        right_frame = LabelFrame(self.root, bd=2, bg='white',
                                 relief=RIDGE, text="Student Records")
        right_frame.place(x=640, y=5, width=600, height=400)

        self.student_table = ttk.Treeview(
            right_frame,
            columns=("ID", "Name", "Division", "Gender",
                     "DOB", "Email", "Phone No",
                     "Address", "Teacher"),
            show="headings"
        )

        for col in self.student_table["columns"]:
            self.student_table.heading(col, text=col)
            self.student_table.column(col, width=90)

        self.student_table.pack(fill="both", expand=True)
        self.student_table.bind("<ButtonRelease>", self.get_cursor)

    # ---------------- EXCEL ---------------- #

    def init_excel_file(self):
        if not os.path.exists(self.excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "Name", "Division", "Gender",
                       "DOB", "Email", "Phone No",
                       "Address", "Teacher"])
            wb.save(self.excel_file)

    def add_data(self):
        if not self.id.get().isdigit():
            messagebox.showerror("Error", "Student ID must be numeric")
            return

        wb = load_workbook(self.excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == self.id.get():
                messagebox.showerror("Error", "Student ID already exists")
                return

        ws.append([
            self.id.get(),
            self.name.get(),
            self.division.get(),
            self.gender.get(),
            self.dob.get(),
            self.email.get(),
            self.phone_no.get(),
            self.address.get(),
            self.teacher.get()
        ])

        wb.save(self.excel_file)
        messagebox.showinfo("Success", "Student added successfully")
        self.load_data()

    def update_data(self):
        wb = load_workbook(self.excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == self.id.get():
                row[1].value = self.name.get()
                row[2].value = self.division.get()
                row[3].value = self.gender.get()
                row[4].value = self.dob.get()
                row[5].value = self.email.get()
                row[6].value = self.phone_no.get()
                row[7].value = self.address.get()
                row[8].value = self.teacher.get()
                wb.save(self.excel_file)
                messagebox.showinfo("Success", "Student updated")
                self.load_data()
                return

        messagebox.showerror("Error", "Student ID not found")

    def delete_data(self):
        wb = load_workbook(self.excel_file)
        ws = wb.active

        for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if str(row[0].value) == self.id.get():
                ws.delete_rows(i)
                wb.save(self.excel_file)
                messagebox.showinfo("Success", "Student deleted")
                self.load_data()
                return

        messagebox.showerror("Error", "Student ID not found")

    def reset_data(self):
        self.id.set("")
        self.name.set("")
        self.division.set("")
        self.gender.set("")
        self.dob.set("")
        self.email.set("")
        self.phone_no.set("")
        self.address.set("")
        self.teacher.set("")

    def load_data(self):
        for item in self.student_table.get_children():
            self.student_table.delete(item)

        wb = load_workbook(self.excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            self.student_table.insert("", "end", values=row)

    def get_cursor(self, event):
        selected = self.student_table.focus()
        values = self.student_table.item(selected, "values")

        if values:
            self.id.set(values[0])
            self.name.set(values[1])
            self.division.set(values[2])
            self.gender.set(values[3])
            self.dob.set(values[4])
            self.email.set(values[5])
            self.phone_no.set(values[6])
            self.address.set(values[7])
            self.teacher.set(values[8])

    # ---------------- FACE CAPTURE ---------------- #

    def take_photo_sample(self):

        if not self.id.get().isdigit():
            messagebox.showerror("Error", "Enter valid numeric Student ID")
            return

        face_cascade = cv2.CascadeClassifier(
            cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
        )

        cap = cv2.VideoCapture(0)
        count = 0

        photo_dir = os.path.join("photos", self.id.get())
        os.makedirs(photo_dir, exist_ok=True)

        while count < 50:
            ret, frame = cap.read()
            if not ret:
                break

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.2, 5)

            for (x, y, w, h) in faces:
                face = gray[y:y+h, x:x+w]
                face = cv2.resize(face, (200, 200))

                count += 1
                cv2.imwrite(os.path.join(photo_dir, f"{count}.jpg"), face)

                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 255, 0), 2)

            cv2.imshow("Capturing Faces", frame)

            if cv2.waitKey(1) & 0xFF == ord("q"):
                break

        cap.release()
        cv2.destroyAllWindows()

        messagebox.showinfo("Success", f"{count} face samples captured successfully!")


if __name__ == "__main__":
    root = tk.Tk()
    app = FaceRecognitionSystem(root)
    root.mainloop()
